"""Read PDF / Excel input files into a structured form the agent can reason over."""

import os
import tempfile
from typing import Any, Dict, List

import boto3
import fitz  # PyMuPDF
from openpyxl import load_workbook


def _resolve_uri(file_uri: str) -> str:
    """Download from S3 if needed, return a local path. Each call gets its own tmp dir."""
    if file_uri.startswith("s3://"):
        without = file_uri[5:]
        bucket, _, key = without.partition("/")
        tmp_dir = tempfile.mkdtemp(prefix="agent-poc-read-")
        local = os.path.join(tmp_dir, os.path.basename(key))
        boto3.client("s3").download_file(bucket, key, local)
        return local
    return file_uri


def read_pdf(file_uri: str) -> Dict[str, Any]:
    path = _resolve_uri(file_uri)
    doc = fitz.open(path)
    pages: List[Dict[str, Any]] = []
    for i, page in enumerate(doc):
        blocks = []
        for b in page.get_text("blocks"):
            x0, y0, x1, y1, text, *_ = b
            text = text.strip()
            if text:
                blocks.append({"bbox": [x0, y0, x1, y1], "text": text})
        pages.append({
            "page": i + 1,
            "width": page.rect.width,
            "height": page.rect.height,
            "blocks": blocks,
        })
    doc.close()
    return {
        "local_path": path,
        "type": "pdf",
        "coord_system": "pdf-points-top-left",
        "pages": pages,
    }


def read_excel(file_uri: str) -> Dict[str, Any]:
    path = _resolve_uri(file_uri)
    wb = load_workbook(path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows = []
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                if cell.value is not None:
                    rows.append({"cell": cell.coordinate, "value": str(cell.value)})
        sheets.append({"name": ws.title, "cells": rows})
    return {"local_path": path, "type": "xlsx", "sheets": sheets}


def read_input_file(file_uri: str, file_type: str) -> Dict[str, Any]:
    """Dispatch on file_type. Returns a structured dict with content + the local path used."""
    if file_type == "pdf":
        return read_pdf(file_uri)
    if file_type in ("xlsx", "excel"):
        return read_excel(file_uri)
    raise ValueError(f"Unsupported file_type: {file_type}")
