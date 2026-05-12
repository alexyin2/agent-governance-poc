"""Write the agent's suggestions back into the source file as annotations / comments.

PDF positioning resolves three modes in priority order:

  1. `bbox`        — exact rectangle (typically obtained via get_pdf_text_positions)
  2. `anchor_text` — search the page for this text; pick the first match
  3. `region`      — coarse area on the page ("full_page" / "top_half" / "bottom_half")

If none resolve, the note falls back to a stacked sticky in the page's top-left
corner so the suggestion still surfaces to the reader.
"""

import logging
import os
import tempfile
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional

import boto3
import fitz  # PyMuPDF
from openpyxl import load_workbook
from openpyxl.comments import Comment

log = logging.getLogger(__name__)

OUTPUT_DIR = os.getenv("OUTPUT_DIR", "outputs")


def _timestamp() -> str:
    """UTC timestamp suitable for file names: YYYYMMDDTHHMMSSZ."""
    return datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%SZ")


def _ensure_local(file_uri: str) -> str:
    if file_uri.startswith("s3://"):
        without = file_uri[5:]
        bucket, _, key = without.partition("/")
        tmp_dir = tempfile.mkdtemp(prefix="agent-poc-write-")
        local = os.path.join(tmp_dir, os.path.basename(key))
        boto3.client("s3").download_file(bucket, key, local)
        return local
    return file_uri


def _output_path(src: str) -> str:
    base = os.path.basename(src)
    name, ext = os.path.splitext(base)
    # Timestamp prevents repeat invocations from clobbering each other in
    # outputs/ (or in S3, where overwrites are silent).
    out_name = f"{name}_revised_{_timestamp()}{ext}"
    # Cloud path (S3_BUCKET set): write to tempdir; container cwd is often read-only.
    # _maybe_upload() will pick the file up and PutObject to s3://bucket/outputs/.
    if os.getenv("S3_BUCKET"):
        tmp_dir = tempfile.mkdtemp(prefix="agent-poc-out-")
        return os.path.join(tmp_dir, out_name)
    # Local dev path: write under repo's outputs/ for easy inspection.
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    return os.path.join(OUTPUT_DIR, out_name)


def _maybe_upload(local: str) -> str:
    bucket = os.getenv("S3_BUCKET")
    if not bucket:
        return local
    key = f"outputs/{os.path.basename(local)}"
    boto3.client("s3").upload_file(local, bucket, key)
    return f"s3://{bucket}/{key}"


_REGION_RECTS = {
    "full_page":   lambda w, h: (20.0, 20.0, w - 20.0, 40.0),
    "top_half":    lambda w, h: (20.0, 20.0, w - 20.0, h / 2.0),
    "bottom_half": lambda w, h: (20.0, h / 2.0, w - 20.0, h - 20.0),
}


def _resolve_position(page: fitz.Page, s: Dict[str, Any]) -> Optional[fitz.Rect]:
    """Resolve a suggestion's position with priority bbox > anchor_text > region.

    Returns None when none of the modes are present or yield a hit; the caller
    falls back to a corner sticky in that case.
    """
    bbox = s.get("bbox")
    if bbox and isinstance(bbox, (list, tuple)) and len(bbox) == 4:
        return fitz.Rect(*bbox)

    anchor = s.get("anchor_text")
    if isinstance(anchor, str) and anchor.strip():
        hits = page.search_for(anchor)
        if hits:
            if len(hits) > 1:
                log.warning(
                    "anchor_text matched %d times on page %d (finding_id=%s); using first hit. anchor=%r",
                    len(hits), page.number + 1, s.get("finding_id", "?"), anchor[:60],
                )
            return hits[0]

    region = s.get("region")
    if isinstance(region, str) and region in _REGION_RECTS:
        w, h = page.rect.width, page.rect.height
        return fitz.Rect(*_REGION_RECTS[region](w, h))

    return None


def write_revised_pdf(file_uri: str, suggestions: List[Dict[str, Any]]) -> str:
    """Annotate a PDF and return the revised file's URI (local or s3://).

    Each suggestion requires `page` (1-indexed) and one of `bbox`, `anchor_text`,
    or `region`. See `_resolve_position` for the resolution order.
    """
    local = _ensure_local(file_uri)
    doc = fitz.open(local)
    try:
        fallback_idx = 0
        for s in suggestions:
            page_idx = max(0, int(s.get("page", 1)) - 1)
            if page_idx >= len(doc):
                continue
            page = doc[page_idx]
            body = s.get("text", "")
            rect = _resolve_position(page, s)
            if rect is not None:
                page.add_highlight_annot(rect)
                page.add_text_annot(fitz.Point(rect.x0, rect.y0), body)
            else:
                # No location info — stack notes in the page corner so the
                # suggestion still surfaces (visually clear it lacks anchor).
                page.add_text_annot(fitz.Point(20, 20 + 15 * fallback_idx), body)
                fallback_idx += 1
        out = _output_path(local)
        doc.save(out)
    finally:
        doc.close()
    return _maybe_upload(out)


def _comment_size(text: str) -> tuple[int, int]:
    """Pick a comment box big enough for the text (Excel default 100x150 px is too small).

    Heuristic: ~24 CJK chars per line, ~18px line height. Min 400x200, max 600x600.
    """
    line_chars = 24
    lines = max(1, sum(max(1, -(-len(part) // line_chars)) for part in text.split("\n")))
    width = 480
    height = max(200, min(600, 40 + lines * 22))
    return width, height


def write_revised_xlsx(file_uri: str, suggestions: List[Dict[str, Any]]) -> str:
    """Attach a comment per suggestion. Each needs {sheet: str, cell: str (e.g. 'A1'), text: str}."""
    local = _ensure_local(file_uri)
    wb = load_workbook(local)
    for s in suggestions:
        sheet = s.get("sheet") or wb.sheetnames[0]
        cell = s.get("cell")
        body = s.get("text", "")
        if not cell or sheet not in wb.sheetnames:
            continue
        comment = Comment(body, "AgentReviewer")
        comment.width, comment.height = _comment_size(body)
        wb[sheet][cell].comment = comment
    out = _output_path(local)
    wb.save(out)
    return _maybe_upload(out)


def annotate_file(file_uri: str, file_type: str, suggestions: List[Dict[str, Any]]) -> str:
    """Dispatch to the PDF or Excel writer based on file_type."""
    if file_type == "pdf":
        return write_revised_pdf(file_uri, suggestions)
    if file_type in ("xlsx", "excel"):
        return write_revised_xlsx(file_uri, suggestions)
    raise ValueError(f"Unsupported file_type: {file_type}")
