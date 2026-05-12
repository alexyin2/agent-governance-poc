"""Sheet-structure lookup for Excel annotation.

`load_file` lets the agent SEE the workbook (text + layout via Bedrock document
block), but the visual view doesn't give it a reliable mapping from "the column
visually labelled 評估意見" to the A1 column letter (`G`?  `H`?). When the agent
needs to write comments to a specific column on a structured checklist, it
calls this tool to get the ground-truth header content + column letters.

The design here is deliberately minimal — we return raw cell text per row with
column letters, and let the agent decide which row is the header and which
column holds review opinions. Server-side heuristics (regex matching "意見|備
註|comment", multi-row header guessing) are intentionally avoided because they
break on locale variations and fail silently when wrong. The agent already has
the visual view to confirm structure; this tool just supplies the letter
mapping it can't derive visually.
"""

import json
import os
import tempfile

import boto3
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


# Most checklist sheets are well under this limit; if exceeded we truncate and
# tell the agent there's more (it can request a different sheet or fall back
# to direct cell addressing).
_MAX_ROWS = 200


def _resolve_uri(file_uri: str) -> str:
    """Download from S3 if needed, return a local path."""
    if file_uri.startswith("s3://"):
        bucket, _, key = file_uri[5:].partition("/")
        tmp_dir = tempfile.mkdtemp(prefix="agent-poc-xlsx-inspect-")
        local = os.path.join(tmp_dir, os.path.basename(key))
        boto3.client("s3").download_file(bucket, key, local)
        return local
    return file_uri


def inspect_xlsx_sheet(file_uri: str, sheet: str | None = None) -> str:
    """Return a sheet's structure (cells with A1-style column letters) for precise annotation.

    When to call:
    - You're about to call ``annotate_xlsx`` on a checklist-style sheet and need
      to know which column letter holds 「評估意見 / 意見 / 備註」 etc.
    - You want to confirm the sheet has the structure you think it has before
      placing comments.

    Args:
        file_uri: Same URI used with ``load_file`` / the pre-loaded document.
        sheet: Optional sheet name. Defaults to the active sheet.

    Returns: JSON string with shape::

        {
          "all_sheets": ["...", ...],     // all worksheet names in the workbook
          "sheet": "...",                  // the sheet you inspected
          "dimensions": "A1:H42",          // used range
          "rows": [
            {"row": 1, "cells": [{"col": "A", "text": "..."}, ...]},
            ...
          ],
          "truncated": null | "showing first N of M rows"
        }

    Empty cells are omitted from each row's `cells` list. You decide which row
    is the header and which column is the "opinion" column from the content;
    no server-side guessing.
    """
    local = _resolve_uri(file_uri)
    wb = load_workbook(local, data_only=True)
    all_sheets = wb.sheetnames
    if sheet is None:
        ws = wb.active
    elif sheet in wb.sheetnames:
        ws = wb[sheet]
    else:
        return json.dumps({
            "error": f"sheet {sheet!r} not found",
            "all_sheets": all_sheets,
        }, ensure_ascii=False)

    rows_out = []
    truncated = None
    total = 0
    for r_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        total = r_idx
        if r_idx > _MAX_ROWS:
            continue
        cells = []
        for cell in row:
            if cell.value is None:
                continue
            text = str(cell.value).strip()
            if not text:
                continue
            cells.append({
                "col": get_column_letter(cell.column),
                "text": text,
            })
        if cells:
            rows_out.append({"row": r_idx, "cells": cells})
    if total > _MAX_ROWS:
        truncated = f"showing first {_MAX_ROWS} of {total} rows"

    return json.dumps(
        {
            "all_sheets": all_sheets,
            "sheet": ws.title,
            "dimensions": ws.dimensions,
            "rows": rows_out,
            "truncated": truncated,
        },
        ensure_ascii=False,
    )
